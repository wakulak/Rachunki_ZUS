from openpyxl import load_workbook

wb = load_workbook('rachunki.xlsx')
sheet = wb.active
file = open('Rachunki.txt', 'w')

cena = 37.79

rep = ("GKM", "GKM ", "KMN", "KMN ", "KM", "KM ")

bydgoszcz = ("Bydgoszcz", "92 10205590 0000 0302 9040 0010", "0010")
koszalin = ("Koszalin", "90 10205590 0000 0502 9130 0011", "0011")
jaslo = ("Jasło", "73 10205590 0000 0302 9110 0013", "0013")
zielona = ("Zielona Góra", "66 10205590 0000 0602 9420 0015", "0015")
elblag = ("Elbląg", "69 10205590 0000 0202 9080 0016", "0016")
czestochowa = ("Częstochowa", "12 10205590 0000 0102 9070 0017", "0017")
chrzanow = ("Chrzanów", "52 10205590 0000 0002 9060 0018", "0018")
radom = ("Radom", "33 10205590 0000 0602 9270 0019", "0019")


def main():
    #zus(bydgoszcz)
    # zus(koszalin)
    # zus(jaslo)
    # zus(zielona)
    # zus(elblag)
    # zus(czestochowa)
    # zus(chrzanow)
    # zus(radom)
    if suma_all() != 0:
        print("Suma wszystkich rachunków pobranych z pliku = %.2f złotych"
                   " (%i rachunków x %.2f złotych)" % (suma_all(), suma_all() / cena, cena))
    else:
        print("Brak rachunków w pliku albo pliku nie można otworzyć")
    zamiana()
    wb.save("rachunki.xlsx")
    file.close()
    print(len(lista_spraw(bydgoszcz)))
    print(len(zam_na_10(lista_spraw(bydgoszcz))))
    print(len(sortowanie(zam_na_10(lista_spraw(bydgoszcz)),rep[0])))
    print(len(sortowanie(zam_na_10(lista_spraw(bydgoszcz)), rep[2])))
    print(len(sortowanie(zam_na_10(lista_spraw(bydgoszcz)), rep[5])))


def lista_spraw(miasto):  # Pobiera do lista_all wszyskie sprawy z pliku wystawione
    # przez wskazany ZUS w argumencie (miasto)
    wiersz = 1
    lista_all = []

    for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):
        wiersz += 1
        for cell in row:
            if cell.value[57:61] == miasto[2]:
                lista_all.append(sheet.cell(row=wiersz, column=1).value)
            else:
                break
    return lista_all


def zam_na_10(lista):  # Argument to lista_spraw(miasto), zwraca nieposegregowaną listę spraw z danego
    # ZUS-u, z sygnaturami w formacie np GKM0001/21, lub KM 0001/21
    lista_10 = []

    for i in lista:
        if (i[0:3]) == rep[0] or (i[0:3]) == rep[2]:
            liczba_zer = 11 - len(i)
            liczba_zer_str = "0" * liczba_zer
            x = 4 - len(i)
            i = i[0:3] + liczba_zer_str + i[x:]
            lista_10.append(i)

        elif (i[0:3]) == rep[5]:
            liczba_zer = 11 - len(i)
            liczba_zer_str = "0" * (liczba_zer - 1)
            x = 4 - len(i)
            i = i[0:3] + liczba_zer_str + i[x - 1:]
            lista_10.append(i)
        else:
            print("Błąd nr 1 w zam_na_10")
    return lista_10


def sortowanie(lista, repertorium, ):  # Pierwszy argument (zam_na_10(lista_spraw(miasto))),
    # drugi argument to jedno z repertorium (GKM,KMN,KM, albo rep[0],rep[2],rep[5]),
    # wychodzi lista posegregowanych rosnąco spraw
    # wskazanego jako argument repertorium
    # format:  GKM0001/21, GKM0002/21
    tablica = [[0 for col in range(10000)] for row in range(30)]

    if repertorium == "KM":
        repertorium = "KM "
    for i in lista:
        if (i[0:3]) == repertorium:
            rok_int = int(i[-2:])
            syg_int = int(i[3:7])
            tablica[rok_int][syg_int] = i
        elif (i[0:3]) != repertorium:
            continue
        else:
            print("Błąd nr 1 w sortowanie")

    for j in range(0, 30):
        while 0 in tablica[j]:
            tablica[j].remove(0)  # usuwa z listy zagnieżdżonnej elementy [0]
    while ([]) in tablica:
        tablica.remove([])  # usuwa puste listy zanieżdżone
    if len(tablica) >= 2:
        while len(tablica) != 1:
            tablica[0] = tablica[0] + tablica[1]  # łączy listy zagnieżdżone (pozostają dwie listy zagnieżdżone)
            del tablica[1]  # ksuje niepotrzebą listę zagnieżdżoną,
    if not tablica:
        return tablica
    else:
        return tablica[0]


def zamiana_10_na_syg(lista):  # Argument to lista
    # (sortowanie(zam_na_10(lista_spraw(miasto)),repertorium))
    # wyjście to posegregowana lista spraw z danego repertorium i danego ZUS-u
    # sygnarury spraw w formacie GKM 1/21, GKM 1999/18, KM 1/19, KMN 123/20
    lista_syg = []

    for i in lista:
        rok_int = (i[-2:])
        sygn_int = int(i[3:7])
        sygn_str = str(sygn_int)
        if (i[0:3]) == rep[5]:
            i = "%s%s/%s" % (i[0:3], sygn_str, rok_int)
        elif i[0:3] == rep[0] or rep[2]:
            i = "%s %s/%s" % (i[0:3], sygn_str, rok_int)
        lista_syg.append(i)
    return lista_syg


def tytul_przelewu(lista):  # argument to (sortowanie(zam_na_10(lista_spraw(miasto)),repertorium))
    # czyli chronologiczna lista spraw danego reprotium w formacie XXX0000/00
    # wyjście lista spraw w formacie 1/19, 7-9/20, 8596/20, 808x3,
    lista_syg_short = []
    lista_tmp = []
    for i in lista:
        rok_int = int(i[-2:])
        sygn_int = int(i[3:7])
        if not lista_tmp:
            lista_tmp.append(i)
        else:
            if rok_int == int(lista_tmp[-1][-2:]) and sygn_int - int(lista_tmp[-1][3:7]) == 1:
                if len(lista_tmp) > 1 and lista_tmp[0] == lista_tmp[-1]:
                    lista_syg_short.append(lista_tmp[0:])
                    lista_tmp.clear()
                    lista_tmp.append(i)
                else:
                    lista_tmp.append(i)
            elif rok_int == int(lista_tmp[-1][-2:]) and sygn_int - int(lista_tmp[-1][3:7]) == 0:
                if i == lista_tmp[0]:
                    lista_tmp.append(i)
                else:
                    lista_syg_short.append(lista_tmp[0:-1])
                    del lista_tmp[0:-1]
                    lista_tmp.append(i)
            else:
                lista_syg_short.append(lista_tmp[0:])
                lista_tmp.clear()
                lista_tmp.append(i)
    if len(lista_tmp) > 0:
        lista_syg_short.append(lista_tmp[0:])
        lista_tmp.clear()

    for i in lista_syg_short:
        rok_int = (i[-1][-2:])
        sygn_int = int(i[0][3:7])
        sygn_str = str(sygn_int)
        if (len(i)) == 1:
            i[0] = ("%s/%s" % (sygn_str, rok_int))
            lista_syg_short[lista_syg_short.index(i)] = i[0]
        elif len(i) > 1 and i[0] != i[1]:
            sygn_int_end = int(i[-1][-7:-3])
            sygn_str_end = str(sygn_int_end)
            i[0] = ("%s-%s/%s" % (sygn_str, sygn_str_end, rok_int))
            lista_syg_short[lista_syg_short.index(i)] = i[0]
        elif len(i) > 1 and i[0] == i[-1]:
            i[0] = ("%s/%sx%s" % (sygn_str, rok_int, len(i)))
            lista_syg_short[lista_syg_short.index(i)] = i[0]

    return lista_syg_short


def suma_miasto(miasto):  # Zwraca koszt wszystkich rachunków wystawionych przez wskazany jako argument ZUS
    licznik = 0
    suma = 0

    for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            if cell.value[57:61] == miasto[2]:
                licznik += 1
                suma = +licznik * cena
    return suma


def suma_all():  # Zwraca koszt wszystkich rachunków pobranych z pliku
    licznik = 0
    suma = 0
    for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            if cell.value[57:61] == bydgoszcz[2] or koszalin[2] or jaslo[2] or zielona[2] \
                    or elblag[2] or czestochowa[2] or chrzanow[2] or radom[2]:
                licznik += 1
        suma = +licznik * cena
    return suma


def zamiana():  # dopisuje do kolumnie 10 i 11 sygnarurę sprawy bez spacji
    wiersz = 1

    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1):
        wiersz += 1
        for cell in row:
            if cell.value[0:3] == rep[0]:
                sheet.cell(row=wiersz, column=11).value = cell.value[4:]
                sheet.cell(row=wiersz, column=10).value = rep[0]
            elif cell.value[0:3] == rep[2]:
                sheet.cell(row=wiersz, column=11).value = cell.value[4:]
                sheet.cell(row=wiersz, column=10).value = rep[2]
            elif cell.value[0:3] == rep[5]:
                sheet.cell(row=wiersz, column=11).value = cell.value[3:]
                sheet.cell(row=wiersz, column=10).value = rep[5]
            else:
                print("\nBłąd nr 1 w def zamiana")


def zus(miasto):
    print("Zus %s, nr rachunku: %s" % (miasto[0], miasto[1]))
    for i in (rep[0], rep[4], rep[2]):
        print("\n\nRachunki do spraw z repertorium %s:\n" % i)
        if not (zamiana_10_na_syg(sortowanie(zam_na_10(lista_spraw(miasto)), i))):
            print("Brak rachunków")
        else:
            x = ", ".join((zamiana_10_na_syg(sortowanie(zam_na_10(lista_spraw(miasto)), i))))
            print(x)
            y = ", ".join(tytul_przelewu(sortowanie(zam_na_10(lista_spraw(miasto)), i)))
            print('\nProponowany tytuł przelewu:\n%s: %s' % (i, y))
            print('\nKwota rachunków z repertorium %s = %.2f złotych, tj. %i x %.2f złotych'
                       % (i, (len(zamiana_10_na_syg(sortowanie(zam_na_10(lista_spraw(miasto)), i))) * cena),
                          len(zamiana_10_na_syg(sortowanie(zam_na_10(lista_spraw(miasto)), i))),cena))
    if suma_miasto(miasto) != 0:
        print("\n\nSuma rachunków z ZUS %s = %.2f złotych, tj. %i x %.2f zł" % (
            miasto[0], suma_miasto(miasto), suma_miasto(miasto) / cena, cena))
    print("\n\n-----------------------------------------------------------------------\n\n")


if __name__ == "__main__":
    main()
