from openpyxl import load_workbook


def main():
    wb = load_workbook('rachunki.xlsx')
    sheet = wb.active

    filename = "rachunki.txt"



    ##Moje zmienne
    cena=37.79
    suma=0

    with open(filename, 'w') as f:
       
        wiersz=1
        licznik=0
        print ("\nZUS Bydgoszcz, nr rachunku: 92 10205590 0000 0302 9040 0010", file=f)
        for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):
            wiersz+=1
            for cell in row:
                if cell.value[57:61] == "0010":
                    print (sheet.cell(row=wiersz, column=1).value,",", end= " ",file=f)
                    licznik=licznik+1
        if licznik>=1:
            print("\nSuma rachunków =",licznik*cena, "złotych, tj.", licznik, "x", cena, "zł",file=f)
        else:
            print ("Brak rachunków",file=f)
        suma+=licznik*cena
            
            
        wiersz=1
        licznik=0
        print ("\nZUS Koszalin, nr rachunku: 90 10205590 0000 0502 9130 0011",file=f)
        for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):
            wiersz+=1
            for cell in row:
                if cell.value[57:61] == "0011":
                    print (sheet.cell(row=wiersz, column=1).value,",", end= " ",file=f)
                    licznik=licznik+1
        if licznik>=1:
            print("\nSuma rachunków =",licznik*cena, "złotych, tj.", licznik, "x", cena, "zł",file=f)
        else:
            print ("Brak rachunków",file=f)
        suma+=licznik*cena
            

        wiersz=1
        licznik=0
        print ("\nZUS Jasło, nr rachunku: 73 10205590 0000 0302 9110 0013",file=f)
        for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):
            wiersz+=1
            for cell in row:
                if cell.value[57:61] == "0013":
                    print (sheet.cell(row=wiersz, column=1).value,",", end= " ",file=f)
                    licznik=licznik+1
        if licznik>=1:
            print("\nSuma rachunków =",licznik*cena, "złotych, tj.", licznik, "x", cena, "zł",file=f)
        else:
            print ("Brak rachunków",file=f)
        suma+=licznik*cena
            

        wiersz=1
        licznik=0
        print ("\nZUS Zielona Góra, nr rachunku: 66 10205590 0000 0602 9420 0015",file=f)
        for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):
            wiersz+=1
            for cell in row:
                if cell.value[57:61] == "0015":
                    print (sheet.cell(row=wiersz, column=1).value,",", end= " ",file=f)
                    licznik=licznik+1
        if licznik>=1:
            print("\nSuma rachunków =",licznik*cena, "złotych, tj.", licznik, "x", cena, "zł",file=f)
        else:
            print ("Brak rachunków",file=f)
        suma+=licznik*cena
            
        wiersz=1
        licznik=0
        print ("\nZUS Elbląg, nr rachunku: 69 10205590 0000 0202 9080 0016",file=f)
        for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):
            wiersz+=1
            for cell in row:
                if cell.value[57:61] == "0016":
                    print (sheet.cell(row=wiersz, column=1).value,",", end= " ",file=f)
                    licznik=licznik+1
        if licznik>=1:
            print("\nSuma rachunków =",licznik*cena, "złotych, tj.", licznik, "x", cena, "zł",file=f)
        else:
            print ("Brak rachunków",file=f)
        suma+=licznik*cena
            
            
        wiersz=1
        licznik=0
        print ("\nZUS Częstochowa, nr rachunku: 12 10205590 0000 0102 9070 0017",file=f)
        for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):
            wiersz+=1
            for cell in row:
                if cell.value[57:61] == "0017":
                    print (sheet.cell(row=wiersz, column=1).value,",", end= " ",file=f)
                    licznik=licznik+1
        if licznik>=1:
            print("\nSuma rachunków =",licznik*cena, "złotych, tj.", licznik, "x", cena, "zł",file=f)
        else:
            print ("Brak rachunków",file=f)
        suma+=licznik*cena
            

        wiersz=1
        licznik=0
        print ("\nZUS Chrzanów, nr rachunku: 52 10205590 0000 0002 9060 0018",file=f)
        for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):
            wiersz+=1
            for cell in row:
                if cell.value[57:61] == "0017":
                    print (sheet.cell(row=wiersz, column=1).value,",", end= " ",file=f)
                    licznik=licznik+1
        if licznik>=1:
            print("\nSuma rachunków =",licznik*cena, "złotych, tj.", licznik, "x", cena, "zł",file=f)
        else:
            print ("Brak rachunków",file=f)
        suma+=licznik*cena
          

        wiersz=1
        licznik=0
        print ("\nZUS Radom, nr rachunku: 33 10205590 0000 0602 9270 0019",file=f)
        for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):
            wiersz+=1
            for cell in row:
                if cell.value[57:61] == "0019":
                    print (sheet.cell(row=wiersz, column=1).value,",", end= " ",file=f)
                    licznik=licznik+1
        if licznik>=1:
            print("\nSuma rachunków =",licznik*cena, "złotych, tj.", licznik, "x", cena, "zł",file=f)
        else:
            print ("Brak rachunków",file=f)
        suma+=licznik*cena

        print("\n\nSuma wszystkich rachunków: ", round(suma,2),file=f)


    f.close()
main()


