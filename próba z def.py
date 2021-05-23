from openpyxl import load_workbook

wb = load_workbook('rachunki.xlsx')
sheet = wb.active
cena=37.79

GKM=("GKM","GKM ")
KMN=("KMN", "KMN ")
KM=("KM ", "KM ")


    
def main():

      bydgoszcz=("Bydgoszcz","92 10205590 0000 0302 9040 0010","0010")
      koszalin=("Koszalin","90 10205590 0000 0502 9130 0011", "0011")
      jaslo=("Jasło","73 10205590 0000 0302 9110 0013", "0013")
      zielona=("Zielona Góra", "66 10205590 0000 0602 9420 0015", "0015")
      elblag=("Elbląg","69 10205590 0000 0202 9080 0016", "0016")
      czestochowa=("Częstochowa","12 10205590 0000 0102 9070 0017","0017")
      chrzanow=("Chrzanów","52 10205590 0000 0002 9060 0018", "0018")
      radom=("Radom", "33 10205590 0000 0602 9270 0019", "0019")

 
      
      zus(bydgoszcz)
      zus(koszalin)
      zus(jaslo)
      zus(zielona)
      zus(elblag)
      zus(czestochowa)
      zus(chrzanow)
      zus(radom)
      suma_all()
      zamiana(GKM,KMN,KM)      
      
      wb.save("rachunki.xlsx")

      
    
   
def zus(miasto):
      lista_GKM=[[0 for col in range(10000)] for row in range(30)]
      lista_KMN=[[0 for col in range(10000)] for row in range(30)]
      lista_KM=[[0 for col in range(10000)] for row in range(30)]
      lista_spraw=[]
      
      wiersz=1
      licznik=0
      lista_spraw=[]

      def sortowanie(lista_IN,rep,rep1,rep2,lista, lista1, lista2):

            def sortowanie_(i,rep,lista): ##przetwarza sygaturę_10 liczbową i dodaje do odp. listy, np. KM 0001/21-> KM 1/21 -> lista_KM[21][1]
                  if (i[0:3])==rep[0]:
                        rok_int=int (i[-2:])                
                        sygn_int=int(i[3:7])              
                        rok_str=str(rok_int)              
                        sygn_str=str(sygn_int)
                        i=rep[1]+sygn_str+"/"+rok_str               
                        lista[rok_int][sygn_int]=i
                        
                  else:
                        print ("Błąd 1")

            def usuwanie_0 (lista): ##tworzy jedną listę prostą posegregowaną sygnaturami
                  for j in range(0,30):
                        while (0) in lista[j]:
                              lista[j].remove(0) ##usuwa z listy zagnieżdżonnej elementy [0]
                  while ([]) in lista:
                        lista.remove([]) ##usuwa puste listy zanieżdżone
            
                  if len(lista)>=2:
                        while len(lista)!=2:
                              lista[0]=lista[0]+lista[1] ##łączy listy zagnieżdżone (pozostają dwie listy zagnieżdżone)
                              del lista[1] ## ksuje niepotrzebą listę zagnieżdżoną, pozostaje lista prosta

                             
                 
                        lista=lista[0]+lista[1]
                        x=", ".join(lista)
                        print ("\n",x)
                       
                        
                        
                  


            def sygn_10_cyf(): ##zamienia wczytaną sygnaturę na sygnaturę 10 cyfrową
                  for i in lista_IN:
                    if(i[0:3])==rep[0]:
                        liczba_zer=11-len(i)
                        liczba_zer_str=("0")*liczba_zer
                        x=4-len(i)
                        i=i[0:3]+liczba_zer_str+i[x:]
                        sortowanie_(i,rep,lista)

                    elif(i[0:3])==rep1[0]:
                        liczba_zer=11-len(i)
                        liczba_zer_str=("0")*liczba_zer
                        x=4-len(i)
                        i=i[0:3]+liczba_zer_str+i[x:]
                        sortowanie_(i,rep1,lista1)
                  
                    elif(i[0:3])==rep2[0]:
                        liczba_zer=11-len(i)
                        liczba_zer_str=("0")*(liczba_zer-1)
                        x=4-len(i)
                        i=i[0:3]+liczba_zer_str+i[x-1:]
                        sortowanie_(i,rep2,lista2)
                        
                    else:
                        print ("Błąd 3")     
                  
        
            sygn_10_cyf()
            usuwanie_0(lista)
            usuwanie_0(lista1)
            usuwanie_0(lista2)
            
            

      
               
      print ("\nZUS", miasto[0], ", nr rachunku:", miasto[1],)
      for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):
        wiersz+=1
        for cell in row:
            if cell.value[57:61] == miasto[2]:
                lista_spraw.append(sheet.cell(row=wiersz, column=1).value)
                licznik=licznik+1
            else:
                break
      if licznik>=1:
        sortowanie (lista_spraw, GKM, KMN, KM,lista_GKM, lista_KMN, lista_KM)      
        print("\nSuma rachunków z ZUS", miasto[0], "=",licznik*cena, "złotych, tj.", licznik, "x", cena, "zł",)
        print ("\n--------------------------------------------------------------------------------")
      else:
        print ("Brak rachunków",)
        print ("\n--------------------------------------------------------------------------------")
      
      

def suma_all():
    licznik=0
    suma=0
    for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            if cell.value[57:61] == "0010" or "0011" or "0013" or "0015" or "0016" or "0017" or "0018" or "0019":
                licznik+=1
            suma=+licznik*cena
    print("\n\nSuma wszystkich rachunków wynosi:", suma, "zl, tj.", licznik, "x", cena, "zł")

def zamiana(rep,rep1,rep2):
    wiersz=1
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1):
        wiersz+=1
        for cell in row:
            if cell.value[0:3] == rep[0]:
                sheet.cell(row=wiersz,column=11).value=cell.value[4:]
                sheet.cell(row=wiersz,column=10).value=rep[0]
            elif cell.value[0:3] == rep1[0]:
                sheet.cell(row=wiersz,column=11).value=cell.value[4:]
                sheet.cell(row=wiersz,column=10).value=rep1[0]
            elif cell.value[0:3] == rep2[0]:
                sheet.cell(row=wiersz,column=11).value=cell.value[3:]
                sheet.cell(row=wiersz,column=10).value=rep2[0]
            else:
                  print ("\nBłąd nr 1 w def zamiana")



   
    
if __name__ == "__main__":
    main()


